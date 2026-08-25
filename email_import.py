#!/usr/bin/env python3
"""
Equipment Received — automated email → Firestore importer.

Runs in GitHub Actions on a schedule. Logs into a Gmail relay inbox over IMAP,
finds the most recent email from the expected sender that has an .xlsx/.xlsm
attachment, parses it with the SAME logic as the website's manual upload, and
writes the arrivals into Firestore. If anything looks wrong, it aborts WITHOUT
writing and exits non-zero so the workflow can flag it (no bad data goes live).

Env vars (provided by the workflow from repo secrets):
  GMAIL_USER            the relay Gmail address (e.g. [email protected])
  GMAIL_APP_PASSWORD    a Gmail App Password (NOT your normal password)
  EXPECTED_SENDER       substring that must appear in the From header. NOTE: because
                        Power Automate re-sends the message, this is the address the
                        FLOW sends from (jareneells@arctic.biz), not Bobby's.
  EXPECTED_SUBJECT      substring that must appear in the Subject (default "Arrivals
                        sheet") — this is the subject set in the Power Automate flow.
  FILENAME_MUST_CONTAIN substring the attachment's filename must contain (default
                        "Equipment Received"). This is what stops OTHER spreadsheets
                        Bobby emails from ever being imported as arrivals.
  PDF_FILENAME_MUST_CONTAIN
                        substring the tool-rental PDF's filename must contain (default "Tool
                        Rental"). NOTE this report is emailed ONCE A MONTH, so on almost every
                        run there is no PDF in the window and "no tool rental PDF" is the
                        normal, healthy outcome — not a fault worth chasing.
  FIREBASE_SA_JSON      full service-account JSON (as a string)
  MAX_AGE_HOURS         optional; only accept an email newer than this (default 26)
  MAX_BODIES            optional; most message bodies to download in one run (default 12).
                        Emails are opened newest-first and we stop once the sheet and the
                        tool PDF are both found, so this only bites on an abnormal inbox —
                        and when it does, the run says so rather than pretending it looked
                        at everything.
  DRY_RUN               optional; "1" parses + validates but does not write
"""

import os, sys, ssl, json, re, imaplib, email, email.utils, hashlib, datetime
from email.header import decode_header

# ---- third-party (installed in the workflow): openpyxl, google-cloud-firestore ----
import openpyxl
from google.cloud import firestore
from google.oauth2 import service_account


# ---------------------------------------------------------------------------
# 1. Helpers that MUST match the website exactly (normJob, makeId, fmtDateKey)
# ---------------------------------------------------------------------------

def norm_job(j):
    """Mirror of normJob(): trim + uppercase."""
    if j is None:
        return ""
    return str(j).strip().upper()


def make_id(parts):
    """
    Mirror of makeId(): djb2-xor hash over the lowercased "|"-joined parts.
    JS: h=5381; h=((h<<5)+h)^charCode; return "a"+(h>>>0).toString(36)+len.toString(36)
    We reproduce 32-bit unsigned overflow and base36 exactly.
    """
    key = "|".join("" if p is None else str(p) for p in parts).lower()
    # JS strings are UTF-16: charCodeAt() yields code UNITS and .length counts them, so a
    # non-BMP character (an emoji pasted into a description) is two units there and one
    # code point here. Iterating Python characters gave a different hash and a different
    # document ID for the same row. Encode to UTF-16 so both sides see the same sequence.
    units = key.encode("utf-16-le")
    h = 5381
    for i in range(0, len(units), 2):
        cu = units[i] | (units[i + 1] << 8)
        # ((h<<5)+h) ^ charCode, kept in 32-bit space like JS bitwise ops
        h = (((h << 5) + h) & 0xFFFFFFFF) ^ cu
        h &= 0xFFFFFFFF
    return "a" + _base36(h) + _base36(len(units) // 2)


def _base36(n):
    """Match JS Number.prototype.toString(36) for non-negative integers."""
    if n == 0:
        return "0"
    digits = "0123456789abcdefghijklmnopqrstuvwxyz"
    out = ""
    while n > 0:
        out = digits[n % 36] + out
        n //= 36
    return out


MONTHS_LONG = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
               "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}


def fmt_date_key(d):
    """
    Mirror of fmtDateKey(): return YYYY-MM-DD (zero-padded) from a date or string.
    Handles real dates, ISO-ish, M/D/Y, and — importantly — dates people TYPE as text
    such as "Thursday, July 16, 2026" or "July 16, 2026". Those show up in the master
    sheet whenever a row is filled in by hand instead of picked, and silently produced
    dateless arrivals before this was handled. Returns "" if unparseable.

    The rules below are deliberately explicit rather than "try to parse anything". The JS
    side used to fall back to new Date(), which accepted "2026/07/09", "7-9-26" and bare
    numbers that these rules reject — so the same row got a date here and no date there,
    two different document IDs, and imported as a duplicate. Both sides now implement the
    same five rules; contract_check.py fails the build if they drift again.
    """
    if d is None or d == "":
        return ""
    if isinstance(d, (datetime.datetime, datetime.date)):
        return f"{d.year:04d}-{d.month:02d}-{d.day:02d}"
    s = str(d).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if m:
        y = m.group(3)
        if len(y) == 2:
            y = "20" + y
        return f"{y}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    # Written-out: optional weekday, month name, day, year.
    #   "Thursday, July 16, 2026" / "July 16, 2026" / "16 July 2026" / "Jul 16 2026"
    m = re.search(r"([A-Za-z]{3,9})\.?\s+(\d{1,2})(?:st|nd|rd|th)?\s*,?\s*(\d{4})", s)
    if m and m.group(1)[:3].lower() in MONTHS_LONG:
        return f"{int(m.group(3)):04d}-{MONTHS_LONG[m.group(1)[:3].lower()]:02d}-{int(m.group(2)):02d}"
    m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]{3,9})\.?\s*,?\s*(\d{4})", s)
    if m and m.group(2)[:3].lower() in MONTHS_LONG:
        return f"{int(m.group(3)):04d}-{MONTHS_LONG[m.group(2)[:3].lower()]:02d}-{int(m.group(1)):02d}"
    return ""


# ---------------------------------------------------------------------------
# 2. Parse the workbook exactly like parseArrivalSheet()
# ---------------------------------------------------------------------------

def parse_arrival_sheet(ws, sheet_name):
    """Mirror of parseArrivalSheet(): skip rental tabs, find header row, map columns."""
    if re.search(r"rental", sheet_name, re.I):
        return []
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []

    # find the header row within the first 6 rows
    h_idx = -1
    for i in range(min(len(rows), 6)):
        joined = "|".join(str(c).lower() if c is not None else "" for c in rows[i])
        if "date received" in joined or ("job" in joined and "description" in joined):
            h_idx = i
            break
    if h_idx < 0:
        h_idx = 1
    if h_idx >= len(rows):
        return []

    H = [str(c).lower() if c is not None else "" for c in rows[h_idx]]

    def col(*keys):
        for k in keys:
            for i, h in enumerate(H):
                if k in h:
                    return i
        return -1

    cD = col("date received", "received")
    cP = col("p.o", "po", "p o")
    cJ = col("job#", "job #", "job num")
    cN = col("job name")
    cDe = col("description")
    cS = col("supplier")
    cDl = col("delivery")
    cR = col("requested")

    out = []
    for i in range(h_idx + 1, len(rows)):
        row = rows[i]

        def g(x):
            return row[x] if (x >= 0 and x < len(row)) else ""

        desc = str(g(cDe) or "").strip()
        dk = fmt_date_key(g(cD))
        if not desc and not dk:
            continue
        if not desc and not str(g(cJ) or "").strip():
            continue
        out.append({
            "dateReceived": dk,
            "po": str(g(cP) or "").strip(),
            "jobNumber": str(g(cJ) or "").strip(),
            "jobName": str(g(cN) or "").strip(),
            "description": desc,
            "supplier": str(g(cS) or "").strip(),
            "deliveryDate": fmt_date_key(g(cDl)),
            "requestedBy": str(g(cR) or "").strip(),
        })
    return out


def parse_rental_sheet(ws):
    """
    Mirror of parseRentalSheet(): rental tabs are positional (columns A..J), not
    header-mapped. Header row = first of the first 4 rows containing "rental".
    """
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []
    h_idx = 0
    for i in range(min(len(rows), 4)):
        joined = "|".join(str(c).lower() if c is not None else "" for c in rows[i])
        if "rental" in joined:
            h_idx = i
            break

    def cell(row, i):
        return row[i] if i < len(row) and row[i] is not None else ""

    out = []
    for i in range(h_idx + 1, len(rows)):
        r = rows[i]
        rid = str(cell(r, 0) or "").strip()
        jn = str(cell(r, 1) or "").strip()
        eq = str(cell(r, 2) or "").strip()
        if not rid and not jn and not eq:
            continue
        po = str(cell(r, 9) or "").strip()
        m = re.search(r"(\d{2}-\d{4})", po) or re.search(r"(\d{2}-\d{4})", jn)
        status_raw = str(cell(r, 6) or "")
        status = "Returned" if re.search(r"return", status_raw, re.I) else (status_raw.strip() or "Renting")
        out.append({
            "rentalId": rid,
            "jobName": jn,
            "equipment": eq,
            "rate": str(cell(r, 3) or "").strip(),
            "vendor": str(cell(r, 4) or "").strip(),
            "dateRented": fmt_date_key(cell(r, 5)),
            "status": status,
            "dateReturned": fmt_date_key(cell(r, 7)),
            "orderedBy": str(cell(r, 8) or "").strip(),
            "po": po,
            "jobNumber": m.group(1) if m else "",
        })
    return out


def parse_rental_sheet(ws):
    """
    Mirror of parseRentalSheet(): fixed column positions, header row found by
    looking for "rental" in the first 4 rows.
    """
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []
    h_idx = 0
    for i in range(min(len(rows), 4)):
        joined = "|".join(str(c).lower() if c is not None else "" for c in rows[i])
        if "rental" in joined:
            h_idx = i
            break

    out = []
    for i in range(h_idx + 1, len(rows)):
        r = rows[i]

        def g(x):
            v = r[x] if x < len(r) else ""
            return "" if v is None else v

        rid = str(g(0)).strip()
        jn = str(g(1)).strip()
        eq = str(g(2)).strip()
        if not rid and not jn and not eq:
            continue
        po = str(g(9)).strip()
        m = re.search(r"(\d{2}-\d{4})", po) or re.search(r"(\d{2}-\d{4})", jn)
        status_raw = str(g(6))
        out.append({
            "rentalId": rid,
            "jobName": jn,
            "equipment": eq,
            "rate": str(g(3)).strip(),
            "vendor": str(g(4)).strip(),
            "dateRented": fmt_date_key(g(5)),
            "status": "Returned" if re.search(r"return", status_raw, re.I) else (status_raw.strip() or "Renting"),
            "dateReturned": fmt_date_key(g(7)),
            "orderedBy": str(g(8)).strip(),
            "po": po,
            "jobNumber": m.group(1) if m else "",
        })
    return out


JOBHDR = re.compile(r"^(CLOSED\s+)?(\d{2}-\d{4})$")
LINEITEM = re.compile(
    r"^(.*?)\s+(\d{1,2}/\d{1,2}/\d{2,4})\s+(\(blank\)|\d{1,2}/\d{1,2}/\d{2,4})"
    r"\s+(\d+)\s+([\d.]+)\s+\$\s*([\d.,]+|-)\s+\$\s*([\d.,]+|-)$"
)


def _pdf_lines(blob):
    """
    Mirror of parseToolPdf(): reconstruct text lines from word positions the same way
    the website's PDF.js code does (bucket by baseline, sort by x, normalize spaces).
    Returns (lines, page_map, page_count).
    """
    import io
    import pdfplumber
    lines, page_map = [], {}
    pages = 0
    with pdfplumber.open(io.BytesIO(blob)) as pdf:
        pages = len(pdf.pages)
        for pno, page in enumerate(pdf.pages, start=1):
            words = page.extract_words(x_tolerance=1.5, y_tolerance=2.5,
                                       keep_blank_chars=False, use_text_flow=False)
            buckets = []
            for w in words:
                y = w["top"]
                b = next((bk for bk in buckets if abs(bk["y"] - y) <= 2.5), None)
                if not b:
                    b = {"y": y, "items": []}
                    buckets.append(b)
                b["items"].append(w)
            buckets.sort(key=lambda b: b["y"])
            for b in buckets:
                b["items"].sort(key=lambda w: w["x0"])
                line = re.sub(r"\s+", " ", " ".join(w["text"] for w in b["items"])).strip()
                lines.append(line)
                jh = JOBHDR.match(line)
                if jh and jh.group(2) not in page_map:
                    page_map[jh.group(2)] = pno
    return lines, page_map, pages


def _tool_skip(l):
    return (not l.strip() or l.startswith("Webduct Tool Rental")
            or l.startswith("Rates are based") or l.startswith("Row Labels")
            or l.startswith("Page ") or l.startswith("Grand Total")
            or bool(re.search(r"\bTotal\b\s*\$", l)))


def parse_tool_lines(lines):
    """Mirror of parseToolLines()."""
    out = []
    job = name = tool = None
    closed = False
    expect_name = False
    for raw in lines:
        l = raw.strip()
        if _tool_skip(l):
            continue
        jh = JOBHDR.match(l)
        if jh:
            closed = bool(jh.group(1))
            job = jh.group(2)
            expect_name = True
            tool = None
            name = None
            continue
        if expect_name:
            name = l
            expect_name = False
            continue
        mi = LINEITEM.match(l)
        if mi:
            end = "" if mi.group(3) == "(blank)" else fmt_date_key(mi.group(3))
            out.append({
                "jobNumber": job, "jobName": name, "jobClosed": closed,
                "toolType": tool or "", "toolId": mi.group(1).strip(),
                "rentalStarted": fmt_date_key(mi.group(2)), "rentalEnded": end,
                "billingDays": int(mi.group(4)) if mi.group(4).isdigit() else 0,
                "dailyRate": float(mi.group(5)) if mi.group(5) else 0,
                "billingTotal": mi.group(6), "discountedRate": mi.group(7),
                "status": "Returned" if end else "Out",
            })
        else:
            tool = l
    return out


def _looks_like_master(wb):
    """
    Sanity check that this workbook is really the Equipment Received master sheet.
    We require at least one non-rental tab whose first few rows contain the
    arrival header words. Cheap, but it stops an unrelated spreadsheet cold.
    """
    for name in wb.sheetnames:
        if re.search(r"rental", name, re.I):
            continue
        ws = wb[name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 6:
                break
            joined = "|".join(str(c).lower() if c is not None else "" for c in row)
            if "date received" in joined or ("job" in joined and "description" in joined):
                return True
    return False


# ---------------------------------------------------------------------------
# 3. Pull the newest matching attachment from the Gmail relay inbox (IMAP)
# ---------------------------------------------------------------------------

def _clean_filename(raw):
    """Decode an attachment filename and strip encoding artifacts/quotes."""
    if not raw:
        return ""
    try:
        dec = decode_header(raw)[0]
        name = dec[0].decode(dec[1] or "utf-8", errors="replace") if isinstance(dec[0], bytes) else str(dec[0])
    except Exception:
        name = str(raw)
    # RFC2231 can leave things like "utf-8''Name.xlsm" or stray quotes.
    name = re.sub(r"^[A-Za-z0-9\-]+''", "", name)
    name = name.strip().strip("'\"").strip()
    # A long filename gets folded across lines in the MIME header, and the fold survives
    # decoding as a newline + leading space INSIDE the name — the relay really does send
    # "Equipment Received & Rentals\n MASTER.xlsm". That only matched FILENAME_MUST_CONTAIN
    # by luck of where it wrapped; one word earlier and the break lands inside "Equipment
    # Received" and the sheet is silently never found. Collapse interior whitespace.
    name = re.sub(r"\s+", " ", name)
    # ...and sometimes percent-encoding survives ("Equipment%20Received.xlsm").
    if "%" in name:
        try:
            from urllib.parse import unquote
            name = unquote(name)
        except Exception:
            pass
    return name.strip()


_ATTACH_CACHE = None


# Message bodies are the only expensive thing here (the master workbook is ~90 KB, and it
# is re-sent in full every time). We walk newest-first, so the sheet we want is almost
# always the first body opened. This cap is the backstop for the day that isn't true: a
# relay loop once left hundreds of copies of the same workbook in the inbox and every poll
# re-walked all of them, which is what pushed a 10-second job past a 10-minute timeout.
MAX_BODIES = int(os.environ.get("MAX_BODIES", "12"))

_IMAP_MON = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Each tuple in a batched FETCH response starts "<seq> (BODY[...". Responses do not come
# back in request order, so the sequence number is how a header gets matched to its message.
_FETCH_SEQ = re.compile(rb"^\s*(\d+)\s+\(")


def _imap_date(d):
    """IMAP wants DD-Mon-YYYY with an English month, whatever the runner's locale is."""
    return f"{d.day:02d}-{_IMAP_MON[d.month - 1]}-{d.year}"


def _imap_quote(s):
    """Quote a value for an IMAP search term."""
    return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'


def _search_ids(imap, since, expected_sender, expected_subject):
    """
    Let the SERVER do the filtering. Gmail can match FROM and SUBJECT itself, which turns
    "every message in the window" into "only the ones that could be the relay".

    Falls back to a date-only search if the server won't take the narrower query — a slow
    correct search beats a fast empty one, and the caller re-checks every header anyway.
    """
    terms = [f"SINCE {since}"]
    if expected_sender:
        terms.append(f"FROM {_imap_quote(expected_sender)}")
    if expected_subject:
        terms.append(f"SUBJECT {_imap_quote(expected_subject)}")
    narrow = "(" + " ".join(terms) + ")"
    if len(terms) > 1:
        try:
            status, data = imap.search(None, narrow)
            if status == "OK":
                return data[0].split(), "sender+subject+date"
            print(f"      (server search returned {status} — falling back to date only)")
        except Exception as e:
            print(f"      (server search rejected: {e} — falling back to date only)")
    status, data = imap.search(None, f"(SINCE {since})")
    if status != "OK":
        raise RuntimeError("IMAP search failed")
    return data[0].split(), "date only"


def _fetch_headers(imap, ids, chunk=300):
    """
    Headers for many messages in ONE fetch rather than a round trip each.

    This is the difference between a job that finishes in seconds and one that times out:
    the old loop paid a network round trip per message in the window, so an inbox holding a
    few hundred relay copies cost a few hundred sequential round trips before any real work.
    """
    out = {}
    for i in range(0, len(ids), chunk):
        status, data = imap.fetch(b",".join(ids[i:i + chunk]),
                                  "(BODY.PEEK[HEADER.FIELDS (FROM SUBJECT DATE)])")
        if status != "OK" or not data:
            continue
        for item in data:
            if not isinstance(item, tuple) or len(item) < 2 or not item[1]:
                continue
            m = _FETCH_SEQ.match(item[0] or b"")
            if m:
                out[m.group(1)] = email.message_from_bytes(item[1])
    return out


def _fetch_structures(imap, ids, chunk=300):
    """
    BODYSTRUCTURE for many messages in one batched fetch — the MIME shape of each message,
    including attachment filenames, WITHOUT downloading any of the attachments.

    Returns {seq: lowercased raw structure} or None if the server won't answer, in which
    case the caller falls back to opening bodies and looking properly.
    """
    out = {}
    for i in range(0, len(ids), chunk):
        try:
            status, data = imap.fetch(b",".join(ids[i:i + chunk]), "(BODYSTRUCTURE)")
        except Exception as e:
            print(f"      (BODYSTRUCTURE unavailable: {e} — will scan bodies instead)")
            return None
        if status != "OK" or not data:
            return None
        for item in data:
            # Normally one flat bytes line per message, but a server may hand back a tuple
            # when the structure contains a literal (a filename sent as {n}).
            if isinstance(item, tuple):
                blob = b"".join(p for p in item if isinstance(p, (bytes, bytearray)))
            elif isinstance(item, (bytes, bytearray)):
                blob = bytes(item)
            else:
                continue
            m = _FETCH_SEQ.match(blob)
            if m:
                out[m.group(1)] = blob.lower()
    return out


def _structure_has(blob, exts):
    """
    Does this BODYSTRUCTURE mention an attachment with one of these extensions?

    Deliberately matches the EXTENSION only, not FILENAME_MUST_CONTAIN: RFC2231 splits a long
    filename across name*0*/name*1* continuations, and a four-character extension is far less
    likely to land on the seam than a two-word phrase. This only decides which bodies are
    worth opening — the authoritative filename check still happens in _clean_filename once
    the body is actually parsed.
    """
    return any(e.encode() in blob for e in exts)


def _hdr_subject(hdr):
    """Decoded Subject, or "" — encoded-word subjects would otherwise never match."""
    raw = hdr.get("Subject", "")
    if not raw:
        return ""
    dec = decode_header(raw)[0]
    return (dec[0].decode(dec[1] or "utf-8", errors="replace")
            if isinstance(dec[0], bytes) else str(dec[0] or ""))


def fetch_attachments(user, app_password, expected_sender, max_age_hours, expected_subject=""):
    """
    ONE IMAP pass that returns the attachments we need as
    [(msg_date, filename, bytes), ...]. Cached for the life of the process so the
    master-sheet and tool-PDF importers share a single fetch instead of doing two.

    Three things keep this independent of how much mail is in the box:
      1. the server filters on FROM/SUBJECT, so we never see unrelated mail;
      2. every header arrives in one batched fetch, not one round trip per message;
      3. bodies are opened newest-first and we stop as soon as both files are in hand.
    """
    global _ATTACH_CACHE
    if _ATTACH_CACHE is not None:
        return _ATTACH_CACHE

    out = []
    imap = imaplib.IMAP4_SSL("imap.gmail.com", ssl_context=ssl.create_default_context())
    try:
        imap.login(user, app_password)
        imap.select("INBOX")

        cutoff = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(hours=max_age_hours)
        # A day of slack on the search: IMAP SINCE is whole-day and server-local, so asking
        # for exactly the cutoff's date can drop a message that is still inside the hour
        # window. The exact cutoff is applied below, against the real Date header.
        since = _imap_date((cutoff - datetime.timedelta(days=1)).date())

        ids, how = _search_ids(imap, since, expected_sender, expected_subject)
        print(f"  IMAP search ({how}, since {since}) matched {len(ids)} message(s)")
        if not ids:
            _ATTACH_CACHE = []
            return _ATTACH_CACHE

        # Re-check sender/subject/date here even when the server already filtered. It costs
        # nothing (no network) and keeps the accept/reject rule identical on both the fast
        # path and the date-only fallback, so the two can never disagree about an email.
        keep = []
        for msg_id, hdr in _fetch_headers(imap, ids).items():
            if expected_sender and expected_sender.lower() not in str(hdr.get("From", "")).lower():
                continue
            if expected_subject and expected_subject.lower() not in _hdr_subject(hdr).lower():
                continue
            try:
                msg_date = email.utils.parsedate_to_datetime(hdr.get("Date"))
                if msg_date.tzinfo is None:
                    msg_date = msg_date.replace(tzinfo=datetime.timezone.utc)
            except Exception:
                msg_date = datetime.datetime.now(datetime.timezone.utc)
            if msg_date < cutoff:
                continue
            keep.append((msg_id, msg_date))

        keep.sort(key=lambda k: k[1], reverse=True)   # newest first
        print(f"  {len(keep)} from the expected sender/subject inside the last {max_age_hours}h")

        xl_want = os.environ.get("FILENAME_MUST_CONTAIN", "Equipment Received").strip().lower()
        pdf_want = os.environ.get("PDF_FILENAME_MUST_CONTAIN", "Tool Rental").strip().lower()

        # Ask what each message CARRIES before opening any of it. One more batched round
        # trip, no attachment data.
        #
        # This is what makes the tool PDF's absence free. It is emailed once a month, so on
        # the other ~29 days "stop when both files are found" never fires and every run used
        # to open the full MAX_BODIES hunting something nobody sent. When the structures come
        # back and none of them mentions a .pdf, that is a definite answer, not a failed
        # search: stop as soon as the sheet is in hand.
        targets, pdf_possible = keep, True
        structs = _fetch_structures(imap, [m for m, _ in keep]) if keep else {}
        if structs:
            xl_c = [(m, d) for m, d in keep if _structure_has(structs.get(m, b""), (".xlsm", ".xlsx"))]
            pdf_c = [(m, d) for m, d in keep if _structure_has(structs.get(m, b""), (".pdf",))]
            if xl_c:
                merged = dict(xl_c)
                merged.update(dict(pdf_c))
                targets = sorted(merged.items(), key=lambda k: k[1], reverse=True)
                pdf_possible = bool(pdf_c)
                print(f"  Carrying a sheet: {len(xl_c)} | carrying a PDF: {len(pdf_c)}"
                      + ("" if pdf_c else "  (none — the tool PDF is monthly, so this is normal)"))
            else:
                # Structures parsed but no spreadsheet in any of them. Could be a filename
                # folded across an RFC2231 continuation, so don't trust it — open bodies.
                print("  No .xlsm/.xlsx seen in the message structures — scanning bodies instead")

        got_xl = got_pdf = False
        opened = 0

        for msg_id, msg_date in targets:
            if got_xl and (got_pdf or not pdf_possible):
                break
            if opened >= MAX_BODIES:
                # Never silent: a cap that isn't reported reads as "found everything".
                missing = [w for w, ok in (("master sheet", got_xl), ("tool PDF", got_pdf)) if not ok]
                print(f"  Stopped at MAX_BODIES={MAX_BODIES} with {len(targets) - opened} "
                      f"message(s) unopened"
                      + (f" — still looking for the {' and '.join(missing)}" if missing else ""))
                break
            status, msg_data = imap.fetch(msg_id, "(RFC822)")
            opened += 1
            if status != "OK" or not msg_data or not msg_data[0]:
                continue
            msg = email.message_from_bytes(msg_data[0][1])
            for part in msg.walk():
                fname = _clean_filename(part.get_filename())
                if not fname:
                    continue
                payload = part.get_payload(decode=True)
                if not payload:
                    continue
                out.append((msg_date, fname, payload))
                low = fname.lower()
                if low.endswith((".xlsx", ".xlsm")) and xl_want in low:
                    got_xl = True
                elif low.endswith(".pdf") and pdf_want in low:
                    got_pdf = True

        print(f"  Opened {opened} message body/bodies")
    finally:
        try:
            imap.logout()
        except Exception:
            pass

    _ATTACH_CACHE = out
    return out


def pick_latest(attachments, filename_must_contain, extensions, what):
    """Newest attachment matching an extension + filename substring."""
    exts = tuple(e.lower() for e in extensions)
    cands = [a for a in attachments
             if a[1].lower().endswith(exts)
             and (not filename_must_contain or filename_must_contain.lower() in a[1].lower())]
    if not cands:
        seen = sorted({a[1] for a in attachments if a[1].lower().endswith(exts)})
        extra = f" Files with the right extension but a non-matching name: {', '.join(seen[:5])}" if seen else ""
        raise RuntimeError(
            f"No email with a {what} attachment ('*{filename_must_contain}*' "
            f"{'/'.join(extensions)}).{extra}")
    cands.sort(key=lambda c: c[0])
    return cands[-1]


# ---------------------------------------------------------------------------
# 4. Main
# ---------------------------------------------------------------------------

def _connect(dry_run):
    if dry_run:
        return None
    sa_info = json.loads(os.environ["FIREBASE_SA_JSON"])
    creds = service_account.Credentials.from_service_account_info(sa_info)
    return firestore.Client(project=sa_info["project_id"], credentials=creds)


class MarkerUnreadable(RuntimeError):
    """The import marker could not be read, so we can't tell if this email is new.

    Raised (not swallowed) so the run FAILS and the workflow opens its once-a-day issue.
    See _already_done for why a read failure must never fall through to a write.
    """


def _already_done(db, marker, email_ms, force):
    """
    True if this exact email was already imported (so repeat polls are cheap); False if not.

    If the marker can't be READ, this RAISES rather than returning — but it still never writes.
    Two things had to be true at once and used to be in tension:

      1. A read failure must not fall through to a write. It used to assume "not imported yet"
         and carry on, which turned a Firestore quota problem into a much worse one: the read
         fails with 429, the importer decides it must import, writes 800+ documents, that fails
         too, the marker is never updated, and the next poll does it all again. So a read
         failure aborts THIS import and writes nothing; the next poll retries cleanly.

      2. But "wrote nothing" must not look like "nothing new." On 2026-08-25 a bad google-api-core
         release made every marker read fail for ~1.5h. The importer skipped safely each time and
         exited 0 — so the job went green, no issue opened, and the sheet simply didn't appear
         until someone noticed by hand. Raising turns that same safe skip into a visible failure,
         and the workflow's failure step opens one deduped issue for the day. Nothing to write is
         still lost, but now it is not silent.

    Skipping is still safe either way: writes go through the same connection, so an import could
    not have succeeded when reads are failing.
    """
    if db is None or force:
        return False
    try:
        prev = db.collection("config").document(marker).get()
    except Exception as e:
        print(f"      !! couldn't read {marker}: {e}")
        print(f"      !! Not writing anything (a read failure must never trigger a blind rewrite).")
        print(f"      !! Flagging this run as failed so it's noticed; the next poll will retry.")
        raise MarkerUnreadable(
            f"couldn't read {marker} ({e}); skipped writing and flagged for a retry") from e
    if prev.exists:
        prev_ms = prev.to_dict().get("emailDateMs")
        if prev_ms and int(prev_ms) == email_ms:
            return True
    return False


def _write_docs(db, coll, docs, base_offset=0):
    base = int(datetime.datetime.now().timestamp() * 1000) - len(docs)
    batch = db.batch()
    n = written = 0
    for i, (doc_id, r) in enumerate(docs.items()):
        payload = dict(r)
        payload["seq"] = base + base_offset + i
        payload["source"] = "email-auto"
        batch.set(db.collection(coll).document(doc_id), payload, merge=True)
        n += 1
        written += 1
        if n >= 400:
            batch.commit()
            batch = db.batch()
            n = 0
    if n:
        batch.commit()
    return written


def import_master_excel(db, atts, dry_run, force, soft_missing):
    """Arrivals + rentals from the Equipment Received & Rentals master workbook."""
    fname_match = os.environ.get("FILENAME_MUST_CONTAIN", "Equipment Received").strip()
    print("\n=== MASTER SHEET (arrivals + rentals) ===")
    print(f"  Looking for '*{fname_match}*' .xlsx/.xlsm ...")
    try:
        msg_date, fname, blob = pick_latest(atts, fname_match, (".xlsx", ".xlsm"), "master sheet")
    except RuntimeError as e:
        if soft_missing:
            print(f"  Nothing to import yet - {e}")
            return False
        raise
    print(f"  Found: {fname}  (sent {msg_date.isoformat()}, {len(blob)} bytes)")

    email_ms = int(msg_date.timestamp() * 1000)
    if _already_done(db, "lastImport", email_ms, force):
        print("  Already imported this exact email - skipping.")
        return False

    import io
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    # CRITICAL: .xlsm files can carry stale dimension metadata, and read-only mode
    # trusts it — silently cutting off rows appended at the bottom (i.e. the NEWEST
    # arrivals). reset_dimensions forces reading to the true end of data.
    for _ws in wb.worksheets:
        try:
            _ws.reset_dimensions()
        except Exception:
            pass
    if not _looks_like_master(wb):
        raise RuntimeError(
            "That workbook doesn't look like the Equipment Received master sheet "
            "(couldn't find the expected 'date received' / 'description' headers). Aborting.")

    # Same split the website does: rental tabs -> rentals, everything else -> arrivals.
    arrivals, rentals = [], []
    for name in wb.sheetnames:
        if re.search(r"rental", name, re.I):
            got = parse_rental_sheet(wb[name])
            rentals.extend(got)
            print(f"    tab '{name}': {len(got)} rentals")
        else:
            got = parse_arrival_sheet(wb[name], name)
            arrivals.extend(got)
            print(f"    tab '{name}': {len(got)} arrivals")
    newest = sorted({a["dateReceived"] for a in arrivals if a["dateReceived"]})[-3:]
    print(f"  Newest arrival dates parsed: {', '.join(newest) if newest else '(none)'}")
    # Loud warning for rows whose date we couldn't read. These still import, but they'd
    # be invisible on the calendar / date sort — exactly the failure that hid July 16.
    undated = [a for a in arrivals if not a["dateReceived"]]
    if undated:
        print(f"  ⚠ {len(undated)} arrival row(s) had an UNREADABLE date and will have no date set:")
        for a in undated[:8]:
            print(f"      job {a['jobNumber'] or '—'} | {a['description'][:44]}")
        print("      (usually means the date was typed as text in the sheet)")

    if not arrivals and not rentals:
        raise RuntimeError("Parsed 0 arrivals and 0 rentals - aborting so no bad data is written.")
    dated = [a for a in arrivals if a["dateReceived"]]
    if arrivals and len(dated) < max(1, int(len(arrivals) * 0.5)):
        raise RuntimeError(
            f"Only {len(dated)}/{len(arrivals)} arrival rows had a valid date - "
            f"the sheet format may have changed. Aborting.")
    print(f"  Parsed {len(arrivals)} arrivals ({len(dated)} dated) | {len(rentals)} rentals")

    a_docs = {make_id([r["dateReceived"], r["po"], norm_job(r["jobNumber"]),
                       r["description"][:80]]): r for r in arrivals}
    r_docs = {make_id([r["rentalId"], norm_job(r["jobNumber"]),
                       r["equipment"][:60], r["dateRented"]]): r for r in rentals}
    print(f"  Unique: {len(a_docs)} arrivals | {len(r_docs)} rentals")

    if dry_run:
        print("  DRY_RUN - not writing. Samples:")
        for k, v in list(a_docs.items())[:3]:
            print(f"    arrival {k} -> {v['dateReceived']} | {v['jobNumber']} | {v['description'][:38]}")
        for k, v in list(r_docs.items())[:3]:
            print(f"    rental  {k} -> {v['dateRented']} | {v['jobNumber']} | {v['equipment'][:38]}")
        return True

    na = _write_docs(db, "arrivals", a_docs)
    nr = _write_docs(db, "rentals", r_docs, base_offset=len(a_docs))
    print(f"  Wrote {na} arrivals | {nr} rentals")
    try:
        db.collection("config").document("lastImport").set({
            "at": int(datetime.datetime.now(datetime.timezone.utc).timestamp() * 1000),
            "emailDate": msg_date.isoformat(), "emailDateMs": email_ms,
            "sourceFile": fname, "count": na, "rentals": nr, "by": "auto (email)",
        }, merge=True)
    except Exception as e:
        print(f"  WARNING: couldn't record lastImport: {e}")
    return True


def import_tool_pdf(db, atts, dry_run, force, soft_missing):
    """Tool rentals from the Webduct Tool Rental PDF (+ store the PDF for the viewer)."""
    pdf_match = os.environ.get("PDF_FILENAME_MUST_CONTAIN", "Tool Rental").strip()
    print("\n=== TOOL RENTAL PDF ===")
    print(f"  Looking for '*{pdf_match}*' .pdf ...")
    try:
        msg_date, fname, blob = pick_latest(atts, pdf_match, (".pdf",), "tool rental PDF")
    except RuntimeError as e:
        if soft_missing:
            print(f"  Nothing to import yet - {e}")
            return False
        raise
    print(f"  Found: {fname}  (sent {msg_date.isoformat()}, {len(blob)} bytes)")

    email_ms = int(msg_date.timestamp() * 1000)
    if _already_done(db, "lastToolImport", email_ms, force):
        print("  Already imported this exact PDF - skipping.")
        return False

    lines, page_map, pages = _pdf_lines(blob)
    items = parse_tool_lines(lines)
    if not items:
        raise RuntimeError(
            "No tool lines found in that PDF - make sure it's the Webduct tool rental "
            "report. Aborting so no bad data is written.")
    jobs = len(set(i["jobNumber"] for i in items))
    print(f"  Parsed {len(items)} tool lines across {jobs} jobs ({pages} pages)")

    t_docs = {make_id([norm_job(r["jobNumber"]), r["toolType"], r["toolId"],
                       r["rentalStarted"]]): r for r in items}
    print(f"  Unique: {len(t_docs)} tool lines")

    if dry_run:
        print("  DRY_RUN - not writing. Samples:")
        for k, v in list(t_docs.items())[:3]:
            print(f"    tool {k} -> {v['jobNumber']} | {v['toolType'][:24]} | {v['toolId']} | {v['status']}")
        return True

    nt = _write_docs(db, "toolRentals", t_docs)
    print(f"  Wrote {nt} tool lines")

    # Store the PDF itself so the in-app "PDF" button can show each job (same as the site).
    try:
        import base64 as _b64
        b64 = _b64.b64encode(blob).decode()
        meta = {"name": fname, "pages": pages, "pageMap": page_map,
                "uploadedAt": firestore.SERVER_TIMESTAMP}
        if len(b64) < 1040000:
            db.collection("pdfStore").document("data").set({"data": b64})
        else:
            meta["tooBig"] = True
            print(f"  NOTE: PDF too large to store ({len(b64)} b64 chars) - viewer will say so.")
        db.collection("pdfStore").document("meta").set(meta)
    except Exception as e:
        print(f"  WARNING: couldn't store PDF: {e}")

    try:
        db.collection("config").document("lastToolImport").set({
            "at": int(datetime.datetime.now(datetime.timezone.utc).timestamp() * 1000),
            "emailDate": msg_date.isoformat(), "emailDateMs": email_ms,
            "sourceFile": fname, "count": nt, "jobs": jobs, "by": "auto (email)",
        }, merge=True)
    except Exception as e:
        print(f"  WARNING: couldn't record lastToolImport: {e}")
    return True


def main():
    user = os.environ["GMAIL_USER"]
    pw = os.environ["GMAIL_APP_PASSWORD"]
    sender = os.environ.get("EXPECTED_SENDER", "").strip()
    subject = os.environ.get("EXPECTED_SUBJECT", "Arrivals sheet").strip()
    max_age = int(os.environ.get("MAX_AGE_HOURS", "26"))
    dry_run = os.environ.get("DRY_RUN", "") == "1"
    force = os.environ.get("FORCE", "") == "1"
    soft_missing = os.environ.get("SOFT_IF_MISSING", "1") == "1"

    # ONE inbox read serves all three imports (arrivals + rentals from the master
    # workbook, tool rentals from the PDF), so polling often stays cheap.
    print(f"Reading relay inbox (from '{sender or 'any'}', subject '{subject or 'any'}', "
          f"last {max_age}h) …")
    atts = fetch_attachments(user, pw, sender, max_age, subject)
    if atts:
        names = sorted({a[1] for a in atts})
        print(f"  {len(atts)} attachment(s) on matching emails: {', '.join(names[:8])}")
    else:
        print("  No matching emails yet.")

    db = _connect(dry_run)
    did_any = False
    errors = []

    # Imported independently — the two files arrive in separate emails, so one being
    # missing or failing must never stop the other.
    for fn in (import_master_excel, import_tool_pdf):
        try:
            if fn(db, atts, dry_run, force, soft_missing):
                did_any = True
        except Exception as e:
            errors.append(f"{fn.__name__}: {e}")
            print(f"  ERROR in {fn.__name__}: {e}", file=sys.stderr)

    print()
    if errors:
        raise RuntimeError(" | ".join(errors))
    if not did_any:
        print("Nothing new to import (this is normal for an early poll).")
    else:
        print("Import complete.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
